import os
import platform
from dataclasses import dataclass
from data.data import data,Execel_Ids
from data.info import info
import openpyxl
from user import user

@dataclass
class tool:
    def clear_screen():
        if platform.system() == "Windows":
            os.system('cls')
        else:
            os.system('clear')

    def verify_extetion_file():
        if data.file_regiter in data.files_for_execel:
            return True
        else:
            return False

    def verify_infos():
        verify_pts = 0
        try:
            if tool.verify_extetion_file():
                verify_pts += 1

            if verify_pts == data.nun_verify:
                return True
            else:
                return False
        except Exception as E:
            print(f"Erro Na Verficaçao, Erro: {E}")

    def set_infos(user_local,ws):
        try:
            saldo = ws[Execel_Ids.Saldo_id].value
            investido = ws[Execel_Ids.Investido_id].value
            despesas = ws[Execel_Ids.Despesas_id].value
            lucro = ws[Execel_Ids.Lucro].value
            saldo_liquido = ws[Execel_Ids.Saldo_Liquido_id].value
            return user(saldo = saldo, investido = investido , lucro = lucro , saldo_liquido = saldo_liquido,despesas= despesas)
        except Exception as E:
            print(f"Erro No Set De Dados!, Erro: {E}")

